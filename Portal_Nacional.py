import os
import time
import datetime
import xml.etree.ElementTree as ET
import pandas as pd
from tqdm import tqdm
import re
import pdfplumber
import calendar
import threading

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys

import customtkinter as ctk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import velopack

URL_PORTAL = "https://www.nfse.gov.br/EmissorNacional"

# Modo global: 'prestados' ou 'tomados'
MODO = 'prestados'

# Defaults baseados no modo
def get_defaults():
    if MODO == 'tomados':
        return r"Z:\01 FISCAL\NFSe\TOMADOS", "11/2025"
    else:
        return r"Z:\01 FISCAL\NFSe\PRESTADOS", "11/2025"

PASTA_DOWNLOADS_DEFAULT, COMPETENCIA_DESEJADA_DEFAULT = get_defaults()
TIMEOUT = 30

# ============================= FUNÇÕES AUXILIARES =============================

def criar_driver(headless=False):
    chrome_options = Options()
    prefs = {
        "download.default_directory": os.path.abspath(PASTA_DOWNLOADS),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "safebrowsing.disable_download_protection": True,
    }
    chrome_options.add_experimental_option("prefs", prefs)
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    if headless:
        chrome_options.add_argument("--headless=new")

    # ✅ SEM Service/ChromeDriverManager (Selenium Manager resolve)
    driver = webdriver.Chrome(options=chrome_options)
    driver.maximize_window()
    return driver

def aguardar_downloads(pasta, timeout=30, log_fn=print):
    log_fn("Aguardando downloads terminarem...")
    start = time.time()
    while time.time() - start < timeout:
        if not any(f.endswith('.crdownload') for f in os.listdir(pasta)):
            time.sleep(2.5)
            if not any(f.endswith('.crdownload') for f in os.listdir(pasta)):
                log_fn("Downloads concluídos.")
                return
        time.sleep(2.5)
    log_fn("Timeout aguardando downloads; seguindo mesmo assim.")

def parse_competencia_str(comp_str):
    try:
        mes, ano = comp_str.split('/')
        return int(ano), int(mes)
    except:
        return None, None

def mesma_competencia(data_str, comp_str):
    try:
        dt = datetime.datetime.strptime(data_str.strip(), "%d/%m/%Y").date()
        ano_c, mes_c = parse_competencia_str(comp_str)
        return dt.year == ano_c and dt.month == mes_c
    except:
        return False

def emissao_anterior_competencia(data_str, comp_str):
    try:
        dt = datetime.datetime.strptime(data_str.strip(), "%d/%m/%Y").date()
        ano_c, mes_c = parse_competencia_str(comp_str)
        return dt.year < ano_c or (dt.year == ano_c and dt.month < mes_c)
    except:
        return False

def obter_situacao_e_numero_da_linha(linha):
    situacao = ""
    numero_nota = ""
    try:
        img = linha.find_element(By.XPATH, ".//img[contains(@src,'tb-cancelada.svg') or contains(@src,'tb-gerada.svg')]")
        src = img.get_attribute("src") or ""
        if "tb-cancelada" in src:
            situacao = "Cancelada"
        elif "tb-gerada" in src:
            situacao = "Autorizada"
    except Exception:
        pass
    try:
        td_num = linha.find_element(By.XPATH, ".//td[contains(@class,'td-numero')]")
        numero_nota = td_num.text.strip()
    except Exception:
        try:
            tds = linha.find_elements(By.TAG_NAME, "td")
            for td in tds:
                txt = td.text.strip()
                if txt.isdigit():
                    numero_nota = txt
                    break
        except Exception:
            pass
    return situacao, numero_nota

def baixar_xml_da_linha(driver, linha, num, comp, situacoes_dict, log_fn):
    try:
        datahora_class = 'td-datahora' if MODO == 'tomados' else 'td-data'
        data_emissao_raw = linha.find_element(By.XPATH, f".//td[contains(@class,'{datahora_class}')]").text.strip()
        data_emissao = data_emissao_raw.split()[0]  # Extrair apenas a parte da data
        # Converter ano de 2 dígitos para 4 dígitos
        partes = data_emissao.split('/')
        if len(partes) == 3 and len(partes[2]) == 2:
            partes[2] = '20' + partes[2]
            data_emissao = '/'.join(partes)
        situacao, numero = obter_situacao_e_numero_da_linha(linha)
        if numero:
            situacoes_dict[numero] = situacao

        if not mesma_competencia(data_emissao, comp):
            log_fn(f"Linha {num}: Ignorada → {data_emissao}")
            if emissao_anterior_competencia(data_emissao, comp):
                return "ANTERIOR"
            return False

        # Clicar no menu apenas para Tomados
        if MODO == 'tomados':
            linha.find_element(By.XPATH, ".//i[contains(@class,'glyphicon-option-vertical')]").click()
            time.sleep(0.5)

        antes_xml = set(os.listdir(PASTA_DOWNLOADS))
        driver.get(linha.find_element(By.XPATH, ".//a[contains(@href,'Download/NFSe/')]").get_attribute("href"))
        aguardar_downloads(PASTA_DOWNLOADS, timeout=0.1)
        novos_xml = [f for f in os.listdir(PASTA_DOWNLOADS) if f.lower().endswith('.xml') and f not in antes_xml]
        if novos_xml:
            SITUACOES_POR_ARQUIVO[novos_xml[0]] = situacao

        try:
            antes_pdf = set(os.listdir(PASTA_DOWNLOADS))
            link_pdf_elt = linha.find_element(By.XPATH, ".//a[contains(@href,'Download/DANFSe/')]")
            driver.get(link_pdf_elt.get_attribute("href"))
            aguardar_downloads(PASTA_DOWNLOADS, timeout=0.5)
            novos_pdf = [f for f in os.listdir(PASTA_DOWNLOADS) if f.lower().endswith('.pdf') and f not in antes_pdf]
            if novos_xml and novos_pdf:
                PDF_POR_ARQUIVO[novos_xml[0]] = novos_pdf[0]
        except Exception as e_pdf:
            log_fn(f"PDF ignorado na linha {num}? Não encontrado: {str(e_pdf)[:80]}")
            pass

        log_fn(f"Linha {num}: BAIXADO → {data_emissao} | {situacao} | Nº {numero}")
        return True
    except Exception as e:
        log_fn(f"Linha {num}: FALHA → {str(e)[:100]}")
        return False

def aplicar_filtro_por_competencia(driver, competencia_str, log_fn=print):
    """
    Preenche os campos:
    - datainicio = 01/MM/AAAA
    - datafim    = último dia do mês (28/29/30/31 conforme mês/ano)
    e clica no botão Filtrar.
    Depois, espera a "mini recarga" terminar.
    """
    mes, ano = competencia_str.split("/")
    mes = int(mes)
    ano = int(ano)

    data_inicio = f"01/{mes:02d}/{ano}"
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    data_fim = f"{ultimo_dia:02d}/{mes:02d}/{ano}"

    log_fn(">>> INICIANDO PREENCHIMENTO DO FILTRO <<<")
    log_fn(f"Aplicando filtro automático: {data_inicio} até {data_fim}")

    inp_inicio = WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.ID, "datainicio"))
    )
    inp_fim = WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_element_located((By.ID, "datafim"))
    )

    def preencher_input(el, valor):
        el.click()
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.DELETE)
        el.send_keys(valor)
        el.send_keys(Keys.TAB)

    preencher_input(inp_inicio, data_inicio)
    preencher_input(inp_fim, data_fim)

    # captura um elemento da tabela atual para detectar "mini-reload"
    try:
        primeira_linha = driver.find_element(By.XPATH, "//table//tbody//tr[td]")
    except:
        primeira_linha = None

    btn_filtrar = WebDriverWait(driver, TIMEOUT).until(
        EC.element_to_be_clickable((By.XPATH, "//button[@type='submit' and contains(.,'Filtrar')]"))
    )
    btn_filtrar.click()

    # espera a tabela antiga ficar stale (sumir/recarregar)
    if primeira_linha:
        WebDriverWait(driver, TIMEOUT).until(EC.staleness_of(primeira_linha))

    # espera a nova tabela (pós-filtro) carregar
    WebDriverWait(driver, TIMEOUT).until(
        EC.presence_of_all_elements_located((By.XPATH, "//table//tbody//tr[td]"))
    )

    log_fn("Filtro aplicado com sucesso.")

def processar_pagina(driver, competencia_str, situacoes_dict, log_fn=print):
    WebDriverWait(driver, TIMEOUT).until(EC.presence_of_all_elements_located((By.XPATH, "//table//tbody//tr[td]")))
    linhas = driver.find_elements(By.XPATH, "//table//tbody//tr[td]")
    log_fn(f"Página atual: {len(linhas)} notas encontradas")
    baixadas = 0
    for i, linha in enumerate(linhas, 1):
        r = baixar_xml_da_linha(driver, linha, i, competencia_str, situacoes_dict, log_fn)
        if r == "ANTERIOR":
            log_fn("Encontrada nota anterior à competência → parando.")
            return -1
        if r is True:
            baixadas += 1
        time.sleep(1.0)
    log_fn(f"→ {baixadas} notas baixadas nesta página")
    return baixadas

def tem_proxima_pagina(driver, log_fn=print):
    pg_param = 'Recebidas' if MODO == 'tomados' else 'Emitidas'
    try:
        btn = driver.find_element(By.XPATH, f"//a[contains(@href, '{pg_param}?pg=') and contains(@data-original-title, 'Próxima')]")
        if "disabled" in btn.find_element(By.XPATH, "./ancestor::li").get_attribute("class"):
            return False
        driver.execute_script("arguments[0].click();", btn)
        time.sleep(2.5)
        return True
    except Exception:
        return False

def safe_float(val):
    try:
        return float(str(val).strip().replace(',', '.'))
    except Exception:
        return 0.0

def get_tag_value(parent, tags, sub_parent=None):
    ns = 'http://www.sped.fazenda.gov.br/nfse'
    if parent is None:
        return 0.0
    base = parent if sub_parent is None else parent.find(f'.//{{{ns}}}{sub_parent}')
    if base is None:
        return 0.0
    for tag_name in tags:
        elem = base.find(f'.//{{{ns}}}{tag_name}')
        if elem is not None and elem.text:
            return safe_float(elem.text.strip())
    return 0.0

def extrair_texto_pdf(pdf_path):
    """Extrai todo o texto de um PDF usando pdfplumber."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                texto_pagina = page.extract_text()
                if texto_pagina:
                    texto_completo += texto_pagina + "\n"
            return texto_completo.strip()
    except Exception as e:
        print(f"Erro ao extrair texto do PDF {pdf_path}: {e}")
        return ""

def parse_dados_nfse_pdf(texto):
    """Parseia dados específicos do texto extraído do PDF de NFS-e."""
    dados = {}

    # Captura textos após os títulos concatenados
    match = re.search(r'SimplesNacionalnaDatadeCompetência\s+RegimedeApuraçãoTributáriapeloSN\s*\n(.*?)\s+(.*?)\s*\n', texto, re.IGNORECASE | re.DOTALL)
    if match:
        simples = match.group(1).strip()
        regime = match.group(2).strip()
        # Formatar textos com substituições manuais
        simples = re.sub(r'Optante-MicroempreendedorIndividual\(MEI\)', 'Optante - Microempreendedor Individual (MEI)', simples)
        simples = re.sub(r'Optante-MicroempresaouEmpresadePequenoPorte\(ME/EPP\)', 'Optante - Microempresa ou Empresa de Pequeno Porte (ME/EPP)', simples)
        simples = re.sub(r'Nãooptante', 'Não optante', simples)
        regime = re.sub(r'RegimedeapuraçãodostributosfederaisemunicipalpeloSimplesNacional', 'Regime de apuração dos tributos federais e municipal pelo Simples Nacional', regime)
        regime = re.sub(r'RegimedeapuraçãodostributosfederaispeloSimplesNacionaleoISSQN', 'Regime de apuração dos tributos federais pelo Simples Nacional eo ISSQN', regime)
        # Adicionar espaços adicionais se necessário
        dados['simples_nacional'] = simples
        dados['regime_apuracao'] = regime

    return dados

def parse_xml_por_nota(xml_path, situacoes_dict=None):
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        ns = '{http://www.sped.fazenda.gov.br/nfse}'

        emit_nome = root.findtext(f'.//{ns}emit/{ns}xNome', '')
        emit_cnpj = root.findtext(f'.//{ns}emit/{ns}CNPJ', '')
        emit_cpf = root.findtext(f'.//{ns}emit/{ns}CPF', '')
        toma_nome = root.findtext(f'.//{ns}toma/{ns}xNome', '')
        toma_cnpj = root.findtext(f'.//{ns}toma/{ns}CNPJ', '')
        toma_cpf = root.findtext(f'.//{ns}toma/{ns}CPF', '')
        n_nfse = root.findtext(f'.//{ns}infNFSe/{ns}nNFSe', '')

        dh_tag = 'dhProc' if MODO == 'tomados' else 'dhEmi'
        dhEmi = root.find(f'.//{ns}{dh_tag}') or root.find(f'.//{ns}dhEmi')
        data_emissao = ''
        if dhEmi is not None and dhEmi.text:
            try:
                data_emissao = datetime.datetime.strptime(dhEmi.text[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
            except:
                pass

        v_bc = safe_float(root.findtext(f'.//{ns}valores/{ns}vBC'))
        v_liq = safe_float(root.findtext(f'.//{ns}valores/{ns}vLiq'))
        v_serv = safe_float(root.findtext(f'.//{ns}valores/{ns}vServ'))
        v_total_ret = safe_float(root.findtext(f'.//{ns}valores/{ns}vTotalRet'))

        v_serv = safe_float(root.findtext(f'.//{ns}infDPS/{ns}valores/{ns}vServPrest/{ns}vServ'))
        x_desc = ''
        codigo_serv = ''
        infDPS = root.find(f'.//{ns}infDPS')
        serv_tag = 'cServ' if MODO == 'tomados' else 'serv'
        serv = infDPS.find(f'.//{ns}{serv_tag}') if infDPS else None
        if serv is not None:
            desc_elt = serv.find(f'.//{ns}xDescServ')
            if desc_elt is not None and desc_elt.text:
                x_desc = desc_elt.text.strip()
            cod_elt = serv.find(f'.//{ns}cTribNac')
            if cod_elt is not None and cod_elt.text:
                codigo_serv = cod_elt.text.strip()

        tribFed = root.find(f'.//{ns}infDPS/{ns}valores/{ns}trib/{ns}tribFed')
        irrf   = get_tag_value(tribFed, ['vRetIRRF'])
        cp     = get_tag_value(tribFed, ['vRetCP'])
        csll   = get_tag_value(tribFed, ['vRetCSLL'])
        pis    = get_tag_value(tribFed, ['vPis'], 'piscofins')
        cofins = get_tag_value(tribFed, ['vCofins'], 'piscofins')

        # === CAMPOS ADICIONAIS ===
        iss_retido = 'N/A'
        valor_iss_retido = 0.0
        tpRet = root.findtext(f'.//{ns}tpRetISSQN', '')
        if tpRet == '2':
            iss_retido = 'SIM'
            valor_iss_retido = safe_float(root.findtext(f'.//{ns}vISSQN', ''))
        elif tpRet == '1':
            iss_retido = 'NÃO'
            valor_iss_retido = 0.0

        emit_documento = emit_cnpj if emit_cnpj else emit_cpf
        toma_documento = toma_cnpj if toma_cnpj else toma_cpf

        nome_arq = os.path.basename(xml_path)
        situacao = SITUACOES_POR_ARQUIVO.get(nome_arq, "Autorizada")
        if situacoes_dict and n_nfse:
            situacao = situacoes_dict.get(n_nfse, situacao)

        return {
            'arquivo': nome_arq,
            'numero_nota': n_nfse,
            'emitente_nome': emit_nome,
            'emitente_cnpj': emit_documento,
            'tomador_nome': toma_nome,
            'tomador_cnpj': toma_documento,
            'data_emissao': data_emissao,
            'valor_bc': v_bc,
            'valor_liq': v_liq,
            'valor_servico': v_serv,
            'descricao_serv': x_desc,
            'codigo_serv': codigo_serv,
            'situacao': situacao,
            'total_retencoes': v_total_ret,
            'irrf': irrf,
            'cp': cp,
            'csll': csll,
            'pis': pis,
            'cofins': cofins,
            'iss_retido': iss_retido,
            'valor_iss_retido': valor_iss_retido
        }
    except Exception as e:
        print(f"[ERRO PARSE] {xml_path}: {e}")
        return None

def limpar_nome_empresa(nome):
    if not nome:
        return "SEM_NOME"
    invalidos = '<>:"/\\|?*'
    for ch in invalidos:
        nome = nome.replace(ch, ' ')
    return ' '.join(nome.split())

def criar_pasta_downloads(pasta):
    os.makedirs(pasta, exist_ok=True)

# Globais
NOTAS_EXISTENTES = set()
SITUACOES_POR_ARQUIVO = {}
PDF_POR_ARQUIVO = {}

def carregar_notas_existentes(pasta_base, competencia_str, log_fn=print):
    global NOTAS_EXISTENTES
    NOTAS_EXISTENTES = set()
    if not os.path.exists(pasta_base):
        return
    tipo = 'tomadas' if MODO == 'tomados' else 'prestadas'
    log_fn(f"Carregando notas {tipo} existentes da competência atual para evitar duplicidade...")
    for root_dir, _, files in os.walk(pasta_base):
        for file in files:
            if file.lower().endswith('.xml'):
                data = parse_xml_por_nota(os.path.join(root_dir, file))
                if data and data['tomador_cnpj' if MODO == 'tomados' else 'emitente_cnpj'] and data['numero_nota'] and data['data_emissao'] and mesma_competencia(data['data_emissao'], competencia_str):
                    NOTAS_EXISTENTES.add((data['tomador_cnpj' if MODO == 'tomados' else 'emitente_cnpj'], data['numero_nota']))
    log_fn(f"Total de notas {tipo} da competência já registradas: {len(NOTAS_EXISTENTES)}")

def gerar_relatorio_para_empresa(pasta_base, pasta_empresa, competencia_str, situacoes_dict, log_fn=print):
    xml_paths = []
    for root_dir, _, files in os.walk(pasta_empresa):
        for f in files:
            if f.lower().endswith('.xml'):
                xml_paths.append(os.path.join(root_dir, f))

    dados = []
    for caminho in tqdm(xml_paths, desc=f"Processando {os.path.basename(pasta_empresa)}"):
        data = parse_xml_por_nota(caminho, situacoes_dict)
        if data and (not data['data_emissao'] or mesma_competencia(data['data_emissao'], competencia_str)):
            # Adicionar dados do PDF para Tomados
            if MODO == 'tomados':
                pdf_dir = os.path.dirname(caminho).replace('XML', 'PDF')
                pdf_nome = f"NFSE N° {data['numero_nota']}.pdf"
                pdf_path = os.path.join(pdf_dir, pdf_nome)
                if os.path.exists(pdf_path):
                    texto_pdf = extrair_texto_pdf(pdf_path)
                    dados_pdf = parse_dados_nfse_pdf(texto_pdf)
                    data['optante_simples'] = dados_pdf.get('simples_nacional', 'N/A')
                    data['regime_apuracao'] = dados_pdf.get('regime_apuracao', 'N/A')
            dados.append(data)

    if not dados:
        log_fn(f"Nenhum dado encontrado em {os.path.basename(pasta_empresa)}")
        return

    df = pd.DataFrame(dados)
    tipo = 'Tomados' if MODO == 'tomados' else 'Prestados'
    df.rename(columns={
        'arquivo': 'Arquivo', 'numero_nota': 'Número da Nota', 'emitente_nome': 'Emitente',
        'emitente_cnpj': 'CNPJ Emitente', 'tomador_nome': 'Tomador', 'tomador_cnpj': 'CNPJ Tomador',
        'data_emissao': 'Data Emissão', 'valor_bc': 'Valor BC', 'valor_liq': 'Valor Líquido',
        'valor_servico': 'Valor Serviço', 'descricao_serv': 'Descrição Serviço', 'codigo_serv': 'Cód. Serviço',
        'situacao': 'Situação', 'total_retencoes': 'Total Retenções',
        'irrf': 'IRRF', 'cp': 'CP', 'csll': 'CSLL', 'pis': 'PIS', 'cofins': 'COFINS',
        'iss_retido': 'ISS RETIDO?', 'valor_iss_retido': 'VALOR DO ISS',
        'optante_simples': 'OPTANTE PELO SIMPLES?', 'regime_apuracao': 'REGIME DE APURAÇÃO'
    }, inplace=True)

    # Reordenar colunas para colocar VALOR DO ISS logo após ISS RETIDO?
    if 'ISS RETIDO?' in df.columns and 'VALOR DO ISS' in df.columns:
        cols = list(df.columns)
        idx = cols.index('ISS RETIDO?')
        if cols[idx + 1] != 'VALOR DO ISS':
            cols.insert(idx + 1, cols.pop(cols.index('VALOR DO ISS')))
        df = df[cols]

    df.sort_values(by=['Data Emissão', 'Número da Nota'], inplace=True, ignore_index=True)
    for col in df.select_dtypes(include='object').columns:
        df[col].replace('', 'N/A', inplace=True)
        df[col].fillna('N/A', inplace=True)

    nome_legivel = os.path.basename(pasta_empresa)
    rel_path = os.path.join(pasta_empresa, f"Relatório {tipo} - {nome_legivel} - {competencia_str.replace('/', '_')}.xlsx")

    with pd.ExcelWriter(rel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Detalhe_Notas', index=False)

    wb = load_workbook(rel_path)
    ws = wb['Detalhe_Notas']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81B3", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 15.43
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 17.25
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.save(rel_path)

    log_fn("="*80)
    log_fn(f"RELATÓRIO GERADO: {nome_legivel}")
    log_fn(f"Total notas: {len(df)} | Líquido: R$ {df['Valor Líquido'].sum():,.2f}")
    log_fn(f"Arquivo: {rel_path}")
    log_fn("="*80)

def organizar_xmls_e_gerar_relatorios_rodada(pasta_base, competencia_str, novos_xmls, situacoes_dict, log_fn=print):
    global NOTAS_EXISTENTES, PDF_POR_ARQUIVO
    empresas = set()
    empresa_nomes = {}
    key_cnpj = 'tomador_cnpj' if MODO == 'tomados' else 'emitente_cnpj'
    key_nome = 'tomador_nome' if MODO == 'tomados' else 'emitente_nome'
    for xml_file in novos_xmls:
        caminho = os.path.join(pasta_base, xml_file)
        data = parse_xml_por_nota(caminho, situacoes_dict)
        if not data:
            try: os.remove(caminho)
            except: pass
            continue

        chave = (data[key_cnpj], data['numero_nota'])
        if chave in NOTAS_EXISTENTES:
            log_fn(f"Duplicado ignorado: {data['numero_nota']}")
            try: os.remove(caminho)
            except: pass
            pdf_assoc = PDF_POR_ARQUIVO.get(xml_file)
            if pdf_assoc:
                pdf_path = os.path.join(pasta_base, pdf_assoc)
                try: os.remove(pdf_path)
                except: pass
            continue
        NOTAS_EXISTENTES.add(chave)

        cnpj_emp = data[key_cnpj]
        if cnpj_emp not in empresa_nomes:
            empresa_nomes[cnpj_emp] = limpar_nome_empresa(data[key_nome])
        nome_pasta = empresa_nomes[cnpj_emp]
        pasta_emp = os.path.join(pasta_base, nome_pasta)
        subpasta = "Canceladas" if data.get('situacao') == "Cancelada" else "Autorizadas"
        dest = os.path.join(pasta_emp, subpasta)
        dest_xml = os.path.join(dest, "XML")
        dest_pdf = os.path.join(dest, "PDF")
        os.makedirs(dest_xml, exist_ok=True)
        os.makedirs(dest_pdf, exist_ok=True)
        os.replace(caminho, os.path.join(dest_xml, xml_file))

        pdf_file = PDF_POR_ARQUIVO.get(xml_file)
        if pdf_file:
            pdf_path = os.path.join(pasta_base, pdf_file)
            if os.path.exists(pdf_path):
                novo_nome = f"NFSE N° {data['numero_nota'] or 'S_N'}.pdf"
                os.replace(pdf_path, os.path.join(dest_pdf, novo_nome))

        empresas.add(pasta_emp)

    for emp in empresas:
        gerar_relatorio_para_empresa(pasta_base, emp, competencia_str, situacoes_dict, log_fn)

# ============================= INTERFACE CUSTOMTKINTER =============================
ctk.set_appearance_mode("system")
ctk.set_default_color_theme("dark-blue")

class NFSeDownloaderApp:
    def __init__(self, root):
        self.font_normal = ctk.CTkFont(family="Segoe UI", size=12)
        self.font_bold = ctk.CTkFont(family="Segoe UI", size=14, weight="bold")
        self.font_title = ctk.CTkFont(family="Segoe UI", size=24, weight="bold")

        self.root = root
        self.root.title("Download NFS-e - Portal Nacional")
        self.root.geometry("1080x720")
        self.root.minsize(1000, 650)

        self.root.after(5000, self.checar_updates_auto)

        # Header
        header = ctk.CTkFrame(root, height=80, corner_radius=0, fg_color="#1e40af")
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(header, text="Portal Nacional - Download de NFS-e",
                     font=self.font_title, text_color="white").pack(pady=20, padx=30, anchor="w")

        main = ctk.CTkFrame(root)
        main.pack(fill="both", expand=True, padx=25, pady=20)

        # Config
        cfg = ctk.CTkFrame(main, corner_radius=12)
        cfg.pack(fill="x", pady=(15, 25))
        ctk.CTkLabel(cfg, text="Configurações", font=self.font_title).pack(pady=20, padx=30, anchor="w")

        # Modo
        r0 = ctk.CTkFrame(cfg)
        r0.pack(fill="x", padx=30, pady=10)
        ctk.CTkLabel(r0, text="Modo:", font=self.font_bold, width=180, anchor="w").pack(side="left", padx=30)
        self.var_modo = ctk.StringVar(value="Prestados")
        ctk.CTkComboBox(r0, values=["Prestados", "Tomados"], variable=self.var_modo, command=self.mudar_modo).pack(side="left", padx=15)

        r1 = ctk.CTkFrame(cfg)
        r1.pack(fill="x", padx=30, pady=10)
        ctk.CTkLabel(r1, text="Pasta de downloads:", font=self.font_bold, width=180, anchor="w").pack(side="left", padx=30)
        self.var_pasta = ctk.StringVar(value=PASTA_DOWNLOADS_DEFAULT)
        ctk.CTkEntry(r1, textvariable=self.var_pasta, height=45, font=self.font_normal).pack(side="left", fill="x", expand=True, padx=(15,0))
        self.btn_procurar = ctk.CTkButton(r1, text="Procurar...", width=130, height=45, font=self.font_bold, fg_color="#2563eb", hover_color="#1d4ed8", command=self.escolher_pasta)
        self.btn_procurar.pack(side="right", padx=(15,0))

        r2 = ctk.CTkFrame(cfg)
        r2.pack(fill="x", padx=30, pady=10)
        ctk.CTkLabel(r2, text="Competência (MM/AAAA):", font=self.font_bold, width=180, anchor="w").pack(side="left", padx=30)
        self.var_comp = ctk.StringVar(value=COMPETENCIA_DESEJADA_DEFAULT)
        ctk.CTkEntry(r2, textvariable=self.var_comp, width=100, height=45, font=self.font_normal, placeholder_text="ex: 11/2025").pack(side="left", padx=15)

        # Botões
        btnspace = ctk.CTkFrame(main, fg_color="transparent")
        btnspace.pack(fill="x", pady=20)
        ctk.CTkButton(btnspace, text="Limpar Log", width=160, height=50, font=self.font_bold, fg_color="#dc2626", hover_color="#b91c1c", command=self.limpar_log).pack(side="left", padx=20)
        self.btn_update = ctk.CTkButton(btnspace, text="Verificar Updates", width=220, height=50, font=self.font_bold, fg_color="#2563eb", hover_color="#1d4ed8", command=self.checar_updates_auto)
        self.btn_update.pack(side="left", padx=12)
        self.btn_start = ctk.CTkButton(btnspace, text="Baixar NFS-e", width=350, height=55,
        font=self.font_bold, fg_color="#1e40af", hover_color="#1d4ed8",
        command=self.iniciar_download)
        self.btn_start.pack(side="right", padx=20)

        # Log
        logbox = ctk.CTkFrame(main, corner_radius=15)
        logbox.pack(fill="both", expand=True, pady=(15,0))
        ctk.CTkLabel(logbox, text="Log da execução", font=self.font_bold).pack(pady=(20,10), padx=30, anchor="w")
        self.txt_log = ctk.CTkTextbox(logbox, font=ctk.CTkFont(family="Consolas", size=12))
        self.txt_log.pack(fill="both", expand=True, padx=30, pady=(0,30))

    def mudar_modo(self, value):
        global MODO, PASTA_DOWNLOADS_DEFAULT, COMPETENCIA_DESEJADA_DEFAULT
        MODO = 'tomados' if value == 'Tomados' else 'prestados'
        PASTA_DOWNLOADS_DEFAULT, COMPETENCIA_DESEJADA_DEFAULT = get_defaults()
        self.var_pasta.set(PASTA_DOWNLOADS_DEFAULT)
        self.var_comp.set(COMPETENCIA_DESEJADA_DEFAULT)
        tipo = 'Tomados' if MODO == 'tomados' else 'Prestados'
        self.btn_start.configure(text=f"Baixar NFS-e {tipo}")

    def log(self, msg):
        self.txt_log.insert("end", msg + "\n")
        self.txt_log.see("end")
        self.root.update_idletasks()
        print(msg)

    def limpar_log(self):
        self.txt_log.delete("0.0", "end")

    def escolher_pasta(self):
        p = filedialog.askdirectory(initialdir=self.var_pasta.get())
        if p:
            self.var_pasta.set(p)
            global PASTA_DOWNLOADS
            PASTA_DOWNLOADS = p

    def iniciar_download(self):
        self.btn_start.configure(state="disabled", text="Processando...")
        thread = threading.Thread(target=self._run_download)
        thread.start()

    def _run_download(self):
        try:
            self._rodar_multiempresas()
        finally:
            self.root.after(0, lambda: self.btn_start.configure(state="normal", text=f"Baixar NFS-e {'Tomados' if MODO == 'tomados' else 'Prestados'}"))

    def _rodar_multiempresas(self):
        global COMPETENCIA_DESEJADA, SITUACOES_POR_ARQUIVO, PDF_POR_ARQUIVO, PASTA_DOWNLOADS
        COMPETENCIA_DESEJADA = self.var_comp.get().strip() or COMPETENCIA_DESEJADA_DEFAULT
        PASTA_DOWNLOADS = self.var_pasta.get().strip() or PASTA_DOWNLOADS_DEFAULT
        tipo = 'tomados' if MODO == 'tomados' else 'prestados'
        secao = 'Tomadas' if MODO == 'tomados' else 'Emitidas'
        self.log("\n" + "="*90)
        self.log(f"Iniciando {tipo} - Competência: {COMPETENCIA_DESEJADA}")
        self.log("="*90)

        carregar_notas_existentes(PASTA_DOWNLOADS, COMPETENCIA_DESEJADA, self.log)
        SITUACOES_POR_ARQUIVO = {}
        PDF_POR_ARQUIVO = {}

        empresa = 0
        while True:
            empresa += 1
            self.log(f"\n{'='*20} EMPRESA #{empresa} {'='*20}")
            driver = None
            try:
                criar_pasta_downloads(PASTA_DOWNLOADS)
                xml_antes = {f for f in os.listdir(PASTA_DOWNLOADS) if f.lower().endswith('.xml')}
                driver = criar_driver(headless=False)
                driver.get(URL_PORTAL)

                # ✅ APLICA O FILTRO ANTES DE QUALQUER DOWNLOAD
                aplicar_filtro_por_competencia(driver, COMPETENCIA_DESEJADA, self.log)

                situacoes_dict = {}
                pagina = 1
                while True:
                    self.log(f"--- PÁGINA {pagina} ---")
                    res = processar_pagina(driver, COMPETENCIA_DESEJADA, situacoes_dict, self.log)
                    aguardar_downloads(PASTA_DOWNLOADS, log_fn=self.log)
                    if res == -1:
                        break
                    if res == 0 and not tem_proxima_pagina(driver, self.log):
                        break
                    if res > 0 and not tem_proxima_pagina(driver, self.log):
                        break
                    pagina += 1

                xml_depois = {f for f in os.listdir(PASTA_DOWNLOADS) if f.lower().endswith('.xml')}
                novos = sorted(xml_depois - xml_antes)
                self.log(f"Novos XMLs nesta empresa: {len(novos)}")
                if novos:
                    organizar_xmls_e_gerar_relatorios_rodada(PASTA_DOWNLOADS, COMPETENCIA_DESEJADA, novos, situacoes_dict, self.log)
            except Exception as e:
                self.log(f"ERRO na empresa {empresa}: SEM MOVIMENTO")
            finally:
                if driver:
                    try:
                        driver.quit()
                    except:
                        pass

            if not messagebox.askyesno("Próxima empresa", "Deseja processar outro CNPJ?"):
                break

        self.log("\n" + "="*90)
        self.log("PROCESSO FINALIZADO COM SUCESSO!")
        self.log("="*90)

    def checar_updates_auto(self):
        try:
            manager = velopack.UpdateManager("https://api.github.com/repos/Pilotto-Contabilidade/Puxar-Notas-PORTAL-NACIONAL/releases")
            self.log("Iniciando verificação de updates...")
            update_info = manager.check_for_updates()
            if update_info:
                self.log(f"Dados de update: {update_info}")
        except Exception as e:
            self.log(f"Erro verificando update: {str(e)}")

# ============================= FINAL =============================
if __name__ == "__main__":
    try:
        velopack.App().run()
    except Exception as e:
        print("Velopack not loaded:", e)
    root = ctk.CTk()
    app = NFSeDownloaderApp(root)
    root.mainloop()
